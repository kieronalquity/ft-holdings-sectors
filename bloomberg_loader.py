"""Bloomberg Excel ingestion and SQLite storage for peer holdings analysis."""

import sqlite3
import re
import logging
from pathlib import Path
from datetime import datetime

import openpyxl
import pandas as pd

from exchange_country_map import get_country_from_exchange, get_country_from_bbg_code

logger = logging.getLogger(__name__)

PEER_SHEETS = ["India", "Asia", "FW"]

# ---------------------------------------------------------------------------
# Database schema
# ---------------------------------------------------------------------------

CREATE_TABLES_SQL = [
    """
    CREATE TABLE IF NOT EXISTS bbg_snapshots (
        snapshot_id    INTEGER PRIMARY KEY AUTOINCREMENT,
        snapshot_date  TEXT NOT NULL,
        file_name      TEXT NOT NULL,
        ingested_at    TEXT NOT NULL,
        snapshot_type  TEXT NOT NULL DEFAULT 'bloomberg',
        UNIQUE(snapshot_date, file_name)
    );
    """,
    """
    CREATE TABLE IF NOT EXISTS bbg_peer_groups (
        id             INTEGER PRIMARY KEY AUTOINCREMENT,
        snapshot_id    INTEGER NOT NULL REFERENCES bbg_snapshots(snapshot_id),
        fund_name      TEXT NOT NULL,
        fund_isin      TEXT,
        peer_set       TEXT NOT NULL,
        is_alquity     INTEGER NOT NULL DEFAULT 0,
        holdings_date  TEXT,
        has_holdings   INTEGER NOT NULL DEFAULT 1,
        sheet_name     TEXT
    );
    """,
    """
    CREATE TABLE IF NOT EXISTS bbg_holdings (
        id             INTEGER PRIMARY KEY AUTOINCREMENT,
        snapshot_id    INTEGER NOT NULL REFERENCES bbg_snapshots(snapshot_id),
        fund_name      TEXT NOT NULL,
        ticker         TEXT NOT NULL,
        weight         REAL NOT NULL,
        exchange_code  TEXT,
        country_derived TEXT,
        is_cash        INTEGER NOT NULL DEFAULT 0
    );
    """,
    """
    CREATE TABLE IF NOT EXISTS bbg_master_data (
        id             INTEGER PRIMARY KEY AUTOINCREMENT,
        snapshot_id    INTEGER NOT NULL REFERENCES bbg_snapshots(snapshot_id),
        ticker         TEXT NOT NULL,
        short_name     TEXT,
        gics_industry  TEXT,
        gics_sector    TEXT,
        country_bbg    TEXT,
        market_cap_raw TEXT,
        market_cap_usd REAL,
        isin           TEXT,
        bb_unique_id   TEXT
    );
    """,
    """
    CREATE TABLE IF NOT EXISTS ft_snapshot_data (
        id             INTEGER PRIMARY KEY AUTOINCREMENT,
        snapshot_id    INTEGER NOT NULL REFERENCES bbg_snapshots(snapshot_id),
        fund_name      TEXT NOT NULL,
        category       TEXT NOT NULL,
        name           TEXT NOT NULL,
        percentage     REAL NOT NULL,
        date_of_data   TEXT,
        peer_set       TEXT NOT NULL DEFAULT '',
        is_alquity     INTEGER NOT NULL DEFAULT 0
    );
    """,
]

CREATE_INDEXES_SQL = [
    "CREATE INDEX IF NOT EXISTS idx_bbg_holdings_snapshot ON bbg_holdings(snapshot_id);",
    "CREATE INDEX IF NOT EXISTS idx_bbg_holdings_fund ON bbg_holdings(snapshot_id, fund_name);",
    "CREATE INDEX IF NOT EXISTS idx_bbg_holdings_ticker ON bbg_holdings(ticker);",
    "CREATE INDEX IF NOT EXISTS idx_bbg_peers_snapshot ON bbg_peer_groups(snapshot_id);",
    "CREATE INDEX IF NOT EXISTS idx_bbg_master_ticker ON bbg_master_data(snapshot_id, ticker);",
    "CREATE INDEX IF NOT EXISTS idx_bbg_master_isin ON bbg_master_data(snapshot_id, isin);",
    "CREATE INDEX IF NOT EXISTS idx_ft_snapshot ON ft_snapshot_data(snapshot_id);",
]


def init_bbg_db(db_path: str) -> None:
    Path(db_path).parent.mkdir(parents=True, exist_ok=True)
    conn = sqlite3.connect(db_path)
    try:
        for sql in CREATE_TABLES_SQL:
            conn.execute(sql)
        for sql in CREATE_INDEXES_SQL:
            conn.execute(sql)
        # Migrate: add snapshot_type column if missing (existing DBs)
        cols = [r[1] for r in conn.execute("PRAGMA table_info(bbg_snapshots)").fetchall()]
        if "snapshot_type" not in cols:
            conn.execute("ALTER TABLE bbg_snapshots ADD COLUMN snapshot_type TEXT NOT NULL DEFAULT 'bloomberg'")
        # Migrate: add peer_set/is_alquity to ft_snapshot_data if missing
        ft_cols = [r[1] for r in conn.execute("PRAGMA table_info(ft_snapshot_data)").fetchall()]
        if ft_cols and "peer_set" not in ft_cols:
            conn.execute("ALTER TABLE ft_snapshot_data ADD COLUMN peer_set TEXT NOT NULL DEFAULT ''")
            conn.execute("ALTER TABLE ft_snapshot_data ADD COLUMN is_alquity INTEGER NOT NULL DEFAULT 0")
        conn.commit()
    finally:
        conn.close()


# ---------------------------------------------------------------------------
# Ingestion
# ---------------------------------------------------------------------------

def ingest_bloomberg_excel(file_path: str, db_path: str, replace: bool = False) -> dict:
    """Ingest a Bloomberg peer holdings Excel file into the database.

    Returns summary dict with counts and any errors.
    """
    path = Path(file_path)
    if not path.exists():
        raise FileNotFoundError(f"File not found: {file_path}")

    init_bbg_db(db_path)

    snapshot_date = _extract_snapshot_date(path.name)
    file_name = path.name

    conn = sqlite3.connect(db_path)
    try:
        # Check for existing snapshot
        existing = conn.execute(
            "SELECT snapshot_id FROM bbg_snapshots WHERE snapshot_date=? AND file_name=?",
            (snapshot_date, file_name),
        ).fetchone()

        if existing:
            if replace:
                sid = existing[0]
                conn.execute("DELETE FROM bbg_holdings WHERE snapshot_id=?", (sid,))
                conn.execute("DELETE FROM bbg_peer_groups WHERE snapshot_id=?", (sid,))
                conn.execute("DELETE FROM bbg_master_data WHERE snapshot_id=?", (sid,))
                conn.execute("DELETE FROM bbg_snapshots WHERE snapshot_id=?", (sid,))
                conn.commit()
            else:
                return {"error": f"Snapshot {snapshot_date} / {file_name} already exists. Use replace=True to overwrite."}

        wb = openpyxl.load_workbook(str(path), data_only=True)

        # Insert snapshot
        cur = conn.execute(
            "INSERT INTO bbg_snapshots (snapshot_date, file_name, ingested_at) VALUES (?, ?, ?)",
            (snapshot_date, file_name, datetime.now().isoformat()),
        )
        snapshot_id = cur.lastrowid

        # Parse peer groups
        no_bbg_isins = _parse_no_bbg_holdings(wb)
        peer_rows, errors = _parse_peer_sheets(wb, snapshot_id, no_bbg_isins)

        # Match fund names to sheet names and parse holdings
        sheet_names = wb.sheetnames
        holdings_rows = []
        for pr in peer_rows:
            if not pr["has_holdings"]:
                continue
            sheet = _match_fund_to_sheet(pr["fund_name"], sheet_names)
            if sheet:
                pr["sheet_name"] = sheet
                fund_holdings = _parse_fund_sheet(wb[sheet], pr["fund_name"])
                holdings_rows.extend(
                    (snapshot_id, pr["fund_name"], h["ticker"], h["weight"],
                     h["exchange_code"], h["country_derived"], h["is_cash"])
                    for h in fund_holdings
                )
            else:
                pr["has_holdings"] = 0
                errors.append(f"No matching sheet for: {pr['fund_name']}")

        # Insert peer groups
        conn.executemany(
            "INSERT INTO bbg_peer_groups "
            "(snapshot_id, fund_name, fund_isin, peer_set, is_alquity, holdings_date, has_holdings, sheet_name) "
            "VALUES (?, ?, ?, ?, ?, ?, ?, ?)",
            [
                (snapshot_id, pr["fund_name"], pr["fund_isin"], pr["peer_set"],
                 pr["is_alquity"], pr["holdings_date"], pr["has_holdings"], pr.get("sheet_name"))
                for pr in peer_rows
            ],
        )

        # Insert holdings
        conn.executemany(
            "INSERT INTO bbg_holdings "
            "(snapshot_id, fund_name, ticker, weight, exchange_code, country_derived, is_cash) "
            "VALUES (?, ?, ?, ?, ?, ?, ?)",
            holdings_rows,
        )

        # Parse and insert master data
        master_rows = _parse_master_data(wb["master_data"], snapshot_id)
        conn.executemany(
            "INSERT INTO bbg_master_data "
            "(snapshot_id, ticker, short_name, gics_industry, gics_sector, "
            "country_bbg, market_cap_raw, market_cap_usd, isin, bb_unique_id) "
            "VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)",
            master_rows,
        )

        conn.commit()

        summary = {
            "snapshot_id": snapshot_id,
            "snapshot_date": snapshot_date,
            "num_funds": len(peer_rows),
            "num_funds_with_holdings": sum(1 for p in peer_rows if p["has_holdings"]),
            "num_holdings": len(holdings_rows),
            "num_master_rows": len(master_rows),
            "errors": errors,
        }
        logger.info("Ingested %s: %s", file_name, summary)
        return summary

    except Exception:
        conn.rollback()
        raise
    finally:
        conn.close()


def _extract_snapshot_date(filename: str) -> str:
    m = re.search(r"(\d{4}-\d{2}-\d{2})", filename)
    if m:
        return m.group(1)
    return datetime.now().strftime("%Y-%m-%d")


def _parse_no_bbg_holdings(wb) -> set:
    """Return set of ISINs from 'No bbg holdings' sheet."""
    isins = set()
    if "No bbg holdings" not in wb.sheetnames:
        return isins
    ws = wb["No bbg holdings"]
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row and row[1]:
            isins.add(str(row[1]).strip())
    return isins


def _parse_peer_sheets(wb, snapshot_id: int, no_bbg_isins: set) -> tuple:
    """Parse India/Asia/FW sheets. Returns (list of peer dicts, list of error strings)."""
    peers = []
    errors = []
    for sheet_name in PEER_SHEETS:
        if sheet_name not in wb.sheetnames:
            errors.append(f"Peer sheet '{sheet_name}' not found")
            continue
        ws = wb[sheet_name]
        for idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True)):
            if not row or not row[0]:
                continue
            name = str(row[0]).strip()
            isin = str(row[1]).strip() if row[1] else None
            holdings_date = None
            if row[3] and not str(row[3]).startswith("#"):
                raw = row[3]
                if isinstance(raw, datetime):
                    holdings_date = raw.strftime("%Y-%m-%d")
                else:
                    holdings_date = str(raw).strip()

            has_holdings = 1
            if isin and isin in no_bbg_isins:
                has_holdings = 0

            peers.append({
                "fund_name": name,
                "fund_isin": isin,
                "peer_set": sheet_name,
                "is_alquity": 1 if idx == 0 else 0,
                "holdings_date": holdings_date,
                "has_holdings": has_holdings,
                "sheet_name": None,
            })
    return peers, errors


def _match_fund_to_sheet(fund_name: str, sheet_names: list) -> str | None:
    """Match a fund name to a potentially truncated Excel sheet name (max 31 chars)."""
    for sn in sheet_names:
        stripped = sn.rstrip()
        if fund_name.startswith(stripped) or stripped.startswith(fund_name[:30]):
            # Verify it's a holdings sheet (not a peer/meta sheet)
            if stripped in ("master_data", "Process", "India", "Asia", "FW", "No bbg holdings"):
                continue
            return sn
    return None


def _parse_fund_sheet(ws, fund_name: str) -> list:
    """Parse a single fund's holdings sheet into list of dicts."""
    holdings = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or not row[0]:
            continue
        ticker = str(row[0]).strip()
        if ticker.startswith("#N/A") or not ticker:
            continue

        weight = 0.0
        if row[1] is not None:
            try:
                weight = float(row[1])
            except (ValueError, TypeError):
                continue

        is_cash = 1 if "Curncy" in ticker else 0
        exchange_code = _extract_exchange_code(ticker)
        country_derived = get_country_from_exchange(exchange_code)

        holdings.append({
            "ticker": ticker,
            "weight": weight,
            "exchange_code": exchange_code,
            "country_derived": country_derived,
            "is_cash": is_cash,
        })
    return holdings


def _extract_exchange_code(ticker: str) -> str | None:
    """Extract exchange code from Bloomberg ticker like 'HDFCB IN Equity' -> 'IN'."""
    parts = ticker.split()
    if len(parts) >= 3 and parts[-1] == "Equity":
        return parts[-2]
    return None


def _parse_master_data(ws, snapshot_id: int) -> list:
    """Parse master_data sheet into list of tuples for insertion."""
    rows = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or not row[0]:
            continue
        ticker = str(row[0]).strip()
        short_name = str(row[1]).strip() if row[1] else None
        gics_industry = str(row[2]).strip() if row[2] else None
        gics_sector = str(row[3]).strip() if row[3] else None
        country_bbg = str(row[4]).strip() if row[4] else None
        market_cap_raw = str(row[5]).strip() if row[5] else None
        market_cap_usd = _parse_market_cap(market_cap_raw)
        isin = str(row[6]).strip() if row[6] else None
        bb_unique_id = str(row[7]).strip() if row[7] else None

        rows.append((
            snapshot_id, ticker, short_name, gics_industry, gics_sector,
            country_bbg, market_cap_raw, market_cap_usd, isin, bb_unique_id,
        ))
    return rows


def _parse_market_cap(raw: str | None) -> float | None:
    """Parse Bloomberg market cap strings to millions. '21.22B' -> 21220.0, '916.81M' -> 916.81."""
    if not raw or raw.startswith("#"):
        return None
    raw = raw.strip()
    try:
        if raw.upper().endswith("T"):
            return float(raw[:-1]) * 1_000_000
        elif raw.upper().endswith("B"):
            return float(raw[:-1]) * 1_000
        elif raw.upper().endswith("M"):
            return float(raw[:-1])
        elif raw.upper().endswith("K"):
            return float(raw[:-1]) / 1_000
        else:
            return float(raw)
    except ValueError:
        return None


# ---------------------------------------------------------------------------
# Query helpers (used by dashboard)
# ---------------------------------------------------------------------------

def get_available_snapshots(db_path: str) -> list:
    conn = sqlite3.connect(db_path)
    try:
        cur = conn.execute(
            "SELECT snapshot_id, snapshot_date, file_name, ingested_at, snapshot_type "
            "FROM bbg_snapshots ORDER BY snapshot_date DESC"
        )
        return [
            {"snapshot_id": r[0], "snapshot_date": r[1], "file_name": r[2],
             "ingested_at": r[3], "snapshot_type": r[4]}
            for r in cur.fetchall()
        ]
    finally:
        conn.close()


def get_peer_funds(db_path: str, snapshot_id: int, peer_set: str) -> pd.DataFrame:
    conn = sqlite3.connect(db_path)
    try:
        return pd.read_sql_query(
            "SELECT * FROM bbg_peer_groups WHERE snapshot_id=? AND peer_set=? ORDER BY is_alquity DESC, fund_name",
            conn, params=(snapshot_id, peer_set),
        )
    finally:
        conn.close()


def load_holdings(db_path: str, snapshot_id: int, peer_set: str,
                  exclude_cash: bool = True, min_weight: float = 0.0) -> pd.DataFrame:
    """Load all holdings for funds in a given peer set, joined with peer group info."""
    conn = sqlite3.connect(db_path)
    try:
        query = """
            SELECT h.fund_name, h.ticker, h.weight, h.exchange_code, h.country_derived, h.is_cash,
                   p.is_alquity, p.peer_set
            FROM bbg_holdings h
            JOIN bbg_peer_groups p ON h.snapshot_id = p.snapshot_id AND h.fund_name = p.fund_name
            WHERE h.snapshot_id = ? AND p.peer_set = ? AND p.has_holdings = 1
        """
        params = [snapshot_id, peer_set]
        if exclude_cash:
            query += " AND h.is_cash = 0"
        if min_weight > 0:
            query += " AND h.weight > ?"
            params.append(min_weight)
        return pd.read_sql_query(query, conn, params=params)
    finally:
        conn.close()


def load_master_data(db_path: str, snapshot_id: int) -> pd.DataFrame:
    conn = sqlite3.connect(db_path)
    try:
        return pd.read_sql_query(
            "SELECT ticker, short_name, gics_industry, gics_sector, country_bbg, "
            "market_cap_usd, isin FROM bbg_master_data WHERE snapshot_id=?",
            conn, params=(snapshot_id,),
        )
    finally:
        conn.close()


def get_latest_bloomberg_snapshot_id(db_path: str) -> int | None:
    """Return the snapshot_id of the most recent bloomberg-type snapshot."""
    conn = sqlite3.connect(db_path)
    try:
        row = conn.execute(
            "SELECT snapshot_id FROM bbg_snapshots WHERE snapshot_type='bloomberg' "
            "ORDER BY snapshot_date DESC LIMIT 1"
        ).fetchone()
        return row[0] if row else None
    finally:
        conn.close()


def create_ft_snapshot(db_path: str, entries: list, label: str | None = None) -> dict:
    """Create an FT-only snapshot from scraped FT entries.

    entries: list of ScrapedEntry objects from scraper.scrape_all_funds().
    label: optional display label, e.g. 'Run 05/03 - only Old view update'.
    """
    from datetime import date as _date

    init_bbg_db(db_path)
    snap_date = _date.today().isoformat()
    file_name = label or f"FT scrape {snap_date}"

    conn = sqlite3.connect(db_path)
    try:
        cur = conn.execute(
            "INSERT INTO bbg_snapshots (snapshot_date, file_name, ingested_at, snapshot_type) "
            "VALUES (?, ?, ?, 'ft_only')",
            (snap_date, file_name, datetime.now().isoformat()),
        )
        snapshot_id = cur.lastrowid

        rows = [
            (snapshot_id, e.fund_name, e.category, e.company_sector,
             e.percentage, e.date_of_data, getattr(e, 'peer_set', ''),
             1 if getattr(e, 'is_alquity', False) else 0)
            for e in entries
        ]
        conn.executemany(
            "INSERT INTO ft_snapshot_data "
            "(snapshot_id, fund_name, category, name, percentage, date_of_data, peer_set, is_alquity) "
            "VALUES (?, ?, ?, ?, ?, ?, ?, ?)",
            rows,
        )
        conn.commit()
        return {"snapshot_id": snapshot_id, "num_entries": len(rows), "label": file_name}
    except Exception:
        conn.rollback()
        raise
    finally:
        conn.close()


# Explicit FT fund name -> Bloomberg fund name mapping (for country data lookup).
# Only funds with a clear match are included; unmatched funds get no country data.
FT_TO_BBG_NAME = {
    # India
    "Alquity SICAV - Alquity Indian Subcontinent Fund USD I Class": "Alquity SICAV-Alquity Indian Subcontinent Y USD",
    "Chikara Indian Subcontinent Fund Class I USD": "Chikara Indian Subcontinent S USD",
    "Goldman Sachs India Equity Portfolio R Inc GBP": "Goldman Sachs India Equity Pf I Acc USD",
    "Jupiter India Select Class D USD Acc": "Jupiter India Select D USD Acc",
    "Kotak Funds - India Midcap Fund A USD Acc": "Kotak Funds-India Midcap C Acc USD",
    "Nomura Funds Ireland - India Equity Fund I Share USD": "Nomura Funds Ireland-India Equity Fund I USD",
    "Stewart Investors Indian Subcontinent All Cap Fund Class B (Accumulation) GBP": "Stewart Investors Indian Sc All Cap Fd B Acc GBP",
    "UTI India Dynamic Equity USD RDR II Dis": "UTI India Dynamic Equity USD Institutional Acc",
    # Asia
    "Alquity SICAV - Alquity Asia Fund USD Y Class": "Alquity SICAV-Alquity Asia Y USD",
    "Matthews Asia Funds - Pacific Tiger Fund I USD Acc": "Matthews Asia Fds-Pacific Tiger I Acc USD",
    "Stewart Investors Asia Pacific Leaders Fund Class B (Accumulation) GBP": "Stewart Investors APac Leaders Fund B Acc GBP",
    # FW
    "Alquity SICAV - Alquity Future World Fund USD Y Class": "Alquity SICAV-Alquity Future World Y USD",
    "Aikya Global Emerging Markets Fund - UCITS I Share Class USD (Accumulating Shares)": "Aikya Global EM Fund - UCITS I USD Voting Acc",
    "Ashmore SICAV Emerging Markets Equity ESG Fund Institutional III USD Acc": "Ashmore SICAV EM Equity ESG Fund I USD Acc",
    "CT (Lux) Responsible Global Emerging Markets Equity Fund I Acc USD": "CT (Lux) Responsible Global Emg Mkts Eq I Acc USD",
    "Candriam Sustainable Equity Emerging Markets C EUR Inc": "Candriam Sustainable Equity Emerging Mkt I EUR C",
    "East Capital Global Emerging Markets Sustainable P EUR": "East Capital Global Emerging Mrkts Sust P EUR Cap",
    "Federated Hermes Global Emerging Markets Ex China Equity Fund F USD Acc": "Federated Hermes Global EM Equity F USD Acc",
    "GQG Partners Emerging Markets Equity Fund Class I US Dollar Accumulating": "GQG Partners Emerging Markets Equity I USD",
    "GemEquity I USD": "GemEquity I USD",
    "Heptagon Fund ICAV - Driehaus Emerging Markets Equity Fund I USD Acc": "Driehaus Emerging Markets Equity I USD",
    "JPMorgan Funds - Emerging Markets Sustainable Equity C Acc USD": "JPM Emerging Markets Sustainable Equity C Acc USD",
    "Matthews Emerging Markets Sustainable Future Fund Institutional Class": "Matthews Emerging Markets Sustainbl Future Fd;Inst",
    "Nordea 1 - Emerging Sustainable Stars Equity Fund BI USD": "Nordea 1 Emerging Stainable Stars Equity BI USD",
    "Polar Capital Funds PLC - Emerging Market Stars Fund I USD Acc": "Polar Capital Emerging Markets Stars I USD Acc",
    "Redwheel Next Generation Emerging Markets Equity Fund I USD Acc": "Redwheel Next Generation EMs Equity I USD",
    "Stewart Investors Global Emerging Markets All Cap Fund Class B (Accumulation) GBP": "Stewart Investors Gl Em Mkts All Cap Fd B Acc GBP",
    "UBAM - Positive Impact Emerging Equity IC USD Accumulation": "UBAM - Positive Impact Emerging Equity IC USD",
    "Vontobel Fund - mtx Emerging Markets Leaders I USD Cap": "Vontobel Fund mtx EM Lds I USD Acc",
}


# Mapping of historical HTML fund names to peer_set and is_alquity
HISTORICAL_FUND_PEER_MAP = {
    # India
    "Chikara Indian Subcontinent Fund Class I USD": ("India", False),
    "Goldman Sachs India Equity Portfolio R Inc GBP": ("India", False),
    "Jupiter India Select Class D USD Acc": ("India", False),
    "Kotak Funds - India Midcap Fund A USD Acc": ("India", False),
    "Stewart Investors Indian Subcontinent All Cap Fund Class B (Accumulation) GBP": ("India", False),
    "UTI India Dynamic Equity USD RDR II Dis": ("India", False),
    # Asia
    "GemAsia I USD": ("Asia", False),
    "Matthews Asia Funds - Pacific Tiger Fund I USD Acc": ("Asia", False),
    "Stewart Investors Asia Pacific Leaders Fund Class B (Accumulation) GBP": ("Asia", False),
    "Veritas Asian Fund Fund D USD": ("Asia", False),
    # FW (all remaining)
    "Aikya Global Emerging Markets Fund - UCITS I Share Class USD (Accumulating Shares)": ("FW", False),
    "Alquity SICAV - VAM Fundamental Emerging Markets Equity Fund EB USD Accumulation": ("FW", False),
    "Ashmore SICAV Emerging Markets Equity ESG Fund Institutional III USD Acc": ("FW", False),
    "Ashmore SICAV Emerging Markets Equity Institutional USD Inc": ("FW", False),
    "Candriam Sustainable Equity Emerging Markets C EUR Inc": ("FW", False),
    "CT (Lux) Responsible Global Emerging Markets Equity Fund I Acc USD": ("FW", False),
    "East Capital Global Emerging Markets Sustainable P EUR": ("FW", False),
    "Federated Hermes Global Emerging Markets Ex China Equity Fund F USD Acc": ("FW", False),
    "FF - Emerging Markets Equity ESG Fund I-Acc-USD": ("FW", False),
    "GemEquity I USD": ("FW", False),
    "GQG Partners Emerging Markets Equity Fund Class I US Dollar Accumulating": ("FW", False),
    "Heptagon Fund ICAV - Driehaus Emerging Markets Equity Fund I USD Acc": ("FW", False),
    "JPMorgan Funds - Emerging Markets Sustainable Equity C Acc USD": ("FW", False),
    "Matthews Emerging Markets Sustainable Future Fund Institutional Class": ("FW", False),
    "Nordea 1 - Emerging Sustainable Stars Equity Fund BI USD": ("FW", False),
    "Polar Capital Funds PLC - Emerging Market Stars Fund I USD Acc": ("FW", False),
    "Redwheel Next Generation Emerging Markets Equity Fund I USD Acc": ("FW", False),
    "Stewart Investors Global Emerging Markets All Cap Fund Class B (Accumulation) GBP": ("FW", False),
    "UBAM - Positive Impact Emerging Equity IC USD Accumulation": ("FW", False),
    "Vontobel Fund - mtx Emerging Markets Leaders I USD Cap": ("FW", False),
}

# Nomura variants (unicode dash vs ?)
for _nomura_name in (
    "Nomura Funds Ireland \u2013 India Equity Fund I Share USD",
    "Nomura Funds Ireland - India Equity Fund I Share USD",
    "Nomura Funds Ireland ? India Equity Fund I Share USD",
):
    HISTORICAL_FUND_PEER_MAP[_nomura_name] = ("India", False)

# Nomura Class A duplicate - also India peer
for _nomura_a in (
    "Nomura Funds Ireland plc - India Equity Fund Class A USD",
    "Nomura Funds Ireland plc \u2013 India Equity Fund Class A USD",
):
    HISTORICAL_FUND_PEER_MAP[_nomura_a] = ("India", False)


def import_historical_html(db_path: str, html_path: str) -> dict:
    """Import historical FT data from the old HTML export.

    Creates one FT snapshot per date column found in the HTML.
    Returns summary dict.
    """
    from bs4 import BeautifulSoup
    from datetime import datetime as _dt

    init_bbg_db(db_path)

    with open(html_path, "r", encoding="utf-8") as f:
        soup = BeautifulSoup(f.read(), "lxml")

    funds = soup.select("li.list-group-item")
    if not funds:
        return {"error": "No funds found in HTML"}

    # Collect all data: {date_str: [(fund_name, category, name, pct, date_of_data, peer_set, is_alquity), ...]}
    date_entries = {}

    for fund_el in funds:
        header_div = fund_el.select_one("div[style*='background-color'] b")
        if not header_div:
            continue
        fund_name = header_div.get_text(strip=True)
        # Normalize unicode dashes
        fund_name = fund_name.replace("\u2013", "-").replace("\u2014", "-").replace("\ufffd", "?")

        peer_set, is_alquity = HISTORICAL_FUND_PEER_MAP.get(fund_name, ("FW", False))

        layout_table = fund_el.select_one("table")
        if not layout_table:
            continue

        tbody = layout_table.find("tbody", recursive=False)
        container = tbody if tbody else layout_table
        trs = container.find_all("tr", recursive=False)

        # trs[0] = Holdings, trs[1] = Sectors, trs[2] = Regions
        categories = ["Holdings", "Sectors", "Regions"]

        for tr_idx, tr in enumerate(trs):
            if tr_idx >= len(categories):
                break
            category = categories[tr_idx]
            tds = tr.find_all("td", recursive=False)

            for td in tds:
                inner_table = td.select_one("table")
                if not inner_table:
                    continue

                rows = inner_table.select("tr")
                if not rows:
                    continue

                # First row is date header
                date_header = rows[0].get_text(strip=True)

                # Second row is "Updated on : DD/MM/YYYY"
                date_of_data = None
                if len(rows) > 1:
                    updated_text = rows[1].get_text(strip=True)
                    import re as _re
                    m = _re.search(r"(\d{2}/\d{2}/\d{4})", updated_text)
                    if m:
                        try:
                            date_of_data = _dt.strptime(m.group(1), "%d/%m/%Y").strftime("%Y-%m-%d")
                        except ValueError:
                            pass

                if date_header not in date_entries:
                    date_entries[date_header] = []

                # Parse data rows (skip header and "Updated on" row)
                count = 0
                for row in rows[2:]:
                    cells = row.select("td")
                    if len(cells) < 3:
                        continue
                    name = cells[1].get_text(strip=True)
                    pct_text = cells[2].get_text(strip=True).replace("%", "").replace(",", "").strip()
                    if not name or not pct_text:
                        continue
                    try:
                        pct = round(float(pct_text), 4)
                    except ValueError:
                        continue

                    date_entries[date_header].append((
                        fund_name, category, name, pct,
                        date_of_data or "", peer_set, 1 if is_alquity else 0,
                    ))
                    count += 1
                    if category == "Holdings" and count >= 10:
                        break

    # Create one FT snapshot per date
    conn = sqlite3.connect(db_path)
    snapshots_created = []
    try:
        for date_header, entries in sorted(date_entries.items()):
            # Parse date_header like "July 29, 2025" to ISO
            try:
                snap_date = _dt.strptime(date_header, "%B %d, %Y").strftime("%Y-%m-%d")
            except ValueError:
                snap_date = date_header

            label = f"FT Historical - {date_header}"

            # Check if already imported
            existing = conn.execute(
                "SELECT snapshot_id FROM bbg_snapshots WHERE file_name=? AND snapshot_type='ft_only'",
                (label,),
            ).fetchone()
            if existing:
                logger.info("Historical snapshot '%s' already exists (id=%d), skipping", label, existing[0])
                snapshots_created.append({"label": label, "snapshot_id": existing[0], "skipped": True})
                continue

            cur = conn.execute(
                "INSERT INTO bbg_snapshots (snapshot_date, file_name, ingested_at, snapshot_type) "
                "VALUES (?, ?, ?, 'ft_only')",
                (snap_date, label, datetime.now().isoformat()),
            )
            snapshot_id = cur.lastrowid

            rows = [
                (snapshot_id, e[0], e[1], e[2], e[3], e[4], e[5], e[6])
                for e in entries
            ]
            conn.executemany(
                "INSERT INTO ft_snapshot_data "
                "(snapshot_id, fund_name, category, name, percentage, date_of_data, peer_set, is_alquity) "
                "VALUES (?, ?, ?, ?, ?, ?, ?, ?)",
                rows,
            )
            snapshots_created.append({
                "label": label, "snapshot_id": snapshot_id,
                "num_entries": len(rows), "snap_date": snap_date,
            })
            logger.info("Imported historical snapshot: %s (%d entries)", label, len(rows))

        conn.commit()
    except Exception:
        conn.rollback()
        raise
    finally:
        conn.close()

    return {"snapshots": snapshots_created, "dates_found": list(date_entries.keys())}


def get_ft_snapshots(db_path: str) -> list:
    """Return all FT-only snapshots ordered by date."""
    conn = sqlite3.connect(db_path)
    try:
        rows = conn.execute(
            "SELECT snapshot_id, snapshot_date, file_name FROM bbg_snapshots "
            "WHERE snapshot_type='ft_only' ORDER BY snapshot_date DESC"
        ).fetchall()
        return [{"snapshot_id": r[0], "snapshot_date": r[1], "file_name": r[2]} for r in rows]
    finally:
        conn.close()


def load_ft_snapshot_data(db_path: str, snapshot_id: int, peer_set: str = "") -> pd.DataFrame:
    """Load FT scraped data for a given snapshot, optionally filtered by peer_set."""
    conn = sqlite3.connect(db_path)
    try:
        if peer_set:
            return pd.read_sql_query(
                "SELECT fund_name, category, name, percentage, date_of_data, is_alquity "
                "FROM ft_snapshot_data WHERE snapshot_id=? AND peer_set=?",
                conn, params=(snapshot_id, peer_set),
            )
        return pd.read_sql_query(
            "SELECT fund_name, category, name, percentage, date_of_data, is_alquity "
            "FROM ft_snapshot_data WHERE snapshot_id=?",
            conn, params=(snapshot_id,),
        )
    finally:
        conn.close()
